from __future__ import annotations

from openpyxl import load_workbook


def carregar_workbook(caminho_arquivo):
    return load_workbook(caminho_arquivo, data_only=True)


def ler_aba_como_dicts(workbook, nome_aba: str) -> list[dict]:
    aba = workbook[nome_aba]
    linhas = list(aba.iter_rows(values_only=True))

    if not linhas:
        return []

    cabecalhos = [str(c).strip() if c is not None else "" for c in linhas[0]]
    registros = []

    for linha in linhas[1:]:
        if linha is None:
            continue

        registro = dict(zip(cabecalhos, linha))

        # ignora linha totalmente vazia
        if all(valor is None or str(valor).strip() == "" for valor in registro.values()):
            continue

        registros.append(registro)

    return registros